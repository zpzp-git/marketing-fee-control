import os
import openpyxl
from pathlib import Path
import re

SOURCE_DIR = Path("docs/source/03 详细设计/3.7页面详细设计")
OUTPUT_DIR = Path("docs/prd/页面详细设计")

def get_merged_cell_value(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # Get the top-left cell of the merged range
            return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return cell.value

def clean_value(val):
    if val is None:
        return ""
    # Convert to string and handle newlines for markdown table
    s = str(val).replace("\n", "<br>")
    # Escape pipe characters for markdown table
    s = s.replace("|", "&#124;")
    return s

def convert_sheet_to_markdown(sheet):
    rows = []
    # Identify images
    # Openpyxl images are stored in sheet._images
    has_images = hasattr(sheet, '_images') and len(sheet._images) > 0
    
    # Get all cells, handling merged ones
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    if max_row == 0 or max_col == 0:
        return "", has_images

    # Pre-process to find empty rows/cols
    grid = []
    for r in range(1, max_row + 1):
        row_data = []
        is_empty = True
        for c in range(1, max_col + 1):
            cell = sheet.cell(row=r, column=c)
            val = get_merged_cell_value(sheet, cell)
            # If it's a formula, we want the formula text if possible
            # Note: with data_only=False, cell.value is the formula string if it starts with '='
            if cell.data_type == 'f':
                val = cell.value
            
            row_data.append(val)
            if val is not None and str(val).strip() != "":
                is_empty = False
        if not is_empty:
            grid.append(row_data)

    if not grid:
        return "", has_images

    # Further clean empty columns
    valid_cols = []
    num_cols = len(grid[0])
    for c in range(num_cols):
        col_has_data = False
        for r in range(len(grid)):
            val = grid[r][c]
            if val is not None and str(val).strip() != "":
                col_has_data = True
                break
        if col_has_data:
            valid_cols.append(c)

    if not valid_cols:
        return "", has_images

    md_lines = []
    # Convert to Markdown Table
    # Use first non-empty row as header
    header = [clean_value(grid[0][i]) for i in valid_cols]
    md_lines.append("| " + " | ".join(header) + " |")
    md_lines.append("| " + " | ".join(["---"] * len(valid_cols)) + " |")
    
    for r in range(1, len(grid)):
        row = [clean_value(grid[r][i]) for i in valid_cols]
        md_lines.append("| " + " | ".join(row) + " |")

    return "\n".join(md_lines), has_images

def process_excel(file_path):
    print(f"Processing: {file_path}")
    try:
        # data_only=False to get formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        rel_path = file_path.relative_to(SOURCE_DIR)
        output_file = OUTPUT_DIR / rel_path.with_suffix(".md")
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        md_content = []
        md_content.append(f"# 原 Excel 文件名：{file_path.name}\n")
        md_content.append(f"- 原始文件路径：`{file_path}`")
        
        sheet_count = len(wb.sheetnames)
        any_images = False
        
        sheets_md = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            table_md, has_images = convert_sheet_to_markdown(sheet)
            if has_images:
                any_images = True
            
            sheets_md.append(f"## Sheet：{sheet_name}\n")
            if has_images:
                sheets_md.append("> **注意：该 Sheet 包含图片，未能结构化识别，待人工确认。**\n")
            
            if table_md:
                sheets_md.append(table_md)
            else:
                sheets_md.append("*（该 Sheet 为空或无有效表格数据）*")
            sheets_md.append("\n")

        md_content.append(f"- 输出文件路径：`{output_file}`")
        md_content.append(f"- Sheet 数量：{sheet_count}")
        md_content.append(f"- 是否检测到图片：{'是' if any_images else '否'}")
        md_content.append("- 转换说明：本文件由 Excel 自动转换，仅保留表格和文字信息，logo、图标、装饰性图片不纳入结构化内容。\n")
        md_content.append("---")
        md_content.extend(sheets_md)
        
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("\n".join(md_content))
            
        return {
            "success": True,
            "source": file_path,
            "dest": output_file,
            "has_images": any_images,
            "error": None
        }
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return {
            "success": False,
            "source": file_path,
            "dest": None,
            "has_images": False,
            "error": str(e)
        }

def main():
    if not SOURCE_DIR.exists():
        print(f"Source directory {SOURCE_DIR} does not exist.")
        return

    excel_files = []
    for root, dirs, files in os.walk(SOURCE_DIR):
        for file in files:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                excel_files.append(Path(root) / file)

    results = []
    for excel_file in excel_files:
        results.append(process_excel(excel_file))

    # Generate Index
    index_file = OUTPUT_DIR / "00-Excel转换索引.md"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    success_count = sum(1 for r in results if r["success"])
    fail_count = len(results) - success_count
    img_files = [r["source"] for r in results if r["has_images"]]
    
    with open(index_file, "w", encoding="utf-8") as f:
        f.write("# Excel 转换 PRD 页面详细设计索引\n\n")
        f.write(f"- 总 Excel 文件数量：{len(results)}\n")
        f.write(f"- 成功转换数量：{success_count}\n")
        f.write(f"- 失败数量：{fail_count}\n\n")
        
        f.write("## 转换详情\n\n")
        f.write("| 源文件 | Markdown 输出路径 | 状态 | 备注 |\n")
        f.write("| --- | --- | --- | --- |\n")
        for r in results:
            status = "成功" if r["success"] else "失败"
            dest = f"`{r['dest']}`" if r["dest"] else "N/A"
            note = r["error"] if r["error"] else ("含图片" if r["has_images"] else "")
            f.write(f"| {r['source'].name} | {dest} | {status} | {note} |\n")
        
        if img_files:
            f.write("\n## 检测到图片但未结构化识别的文件列表\n\n")
            for img_file in img_files:
                f.write(f"- `{img_file}`\n")

    print("\nConversion Finished!")
    print(f"Total Excel Files: {len(results)}")
    print(f"Successfully Converted: {success_count}")
    print(f"Failed: {fail_count}")

if __name__ == "__main__":
    main()
